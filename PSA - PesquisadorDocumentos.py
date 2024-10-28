import os
import glob
import openpyxl
import time
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from ttkbootstrap import Style
from tkinter import ttk
import pandas as pd
import re
from concurrent.futures import ThreadPoolExecutor

diretorio_selecionado = ""

def carregar_documentos(documentos_textbox): #função responsável por pegar os documentos do textbox para realizar a pesquisa dos documentos
    if not documentos_textbox.get("1.0", tk.END).strip():
        messagebox.showerror("Erro", "Por favor, insira os documentos no campo de texto.")  
        return
    
    documentos = documentos_textbox.get("1.0", tk.END).strip().split("\n") #declarando a variável documentos e pegando com quebra de linha os mesmos e armazenando
    
    return [documento.strip() for documento in documentos if documento.strip()] #retornando uma lista de documentos já ordenados

def pesquisar_documentos(diretorio_atual, extensoes_descartadas, validacao_nomenclatura, documentos):
    documentos_encontrados = {}
    documentos_nao_encontrados = set(documentos)
    qtde_arquivos_pesquisados = 0
    
    # Prepara um regex de busca com os documentos
    regex_documentos = re.compile('|'.join(re.escape(doc) for doc in documentos))
    
    # Função para processar cada arquivo
    def processar_arquivo(file_path):
        nonlocal qtde_arquivos_pesquisados
        
        #Abre o arquivo apenas uma vez
        try:
            with open(file_path, errors="ignore") as f:
                conteudo = f.read()
                
                # Verifica se há algum documento no arquivo
                documentos_encontrados_no_arquivo = regex_documentos.findall(conteudo)
                if documentos_encontrados_no_arquivo:
                    for documento in documentos_encontrados_no_arquivo:
                        if documento not in documentos_encontrados:
                            documentos_encontrados[documento] = []
                        documentos_encontrados[documento].append(os.path.basename(file_path))
                        documentos_nao_encontrados.discard(documento)
                        
                qtde_arquivos_pesquisados += 1
        
        except Exception as e:
            print(f"Erro ao ler o arquivo {file_path}: {e}")

    #Loop pelos arquivos no diretório
    with ThreadPoolExecutor() as executor:
        futures = []
        for root, _, files in os.walk(diretorio_atual):
            for file in files:
                # Filtra arquivos por extensões e nome
                if not any(file.endswith(ext) for ext in extensoes_descartadas) and \
                   not any(file.startswith(validacao) for validacao in validacao_nomenclatura):
                    
                    #Adiciona o arquivo para processamento
                    file_path = os.path.join(root, file)
                    futures.append(executor.submit(processar_arquivo, file_path))
        
        #Espera todos os arquivos serem processados
        for future in futures:
            future.result()

    qtde_documentos_nao_encontrados = len(documentos_nao_encontrados)
    qtde_documentos_encontrados = len(documentos_encontrados)
    
    return documentos_encontrados, documentos_nao_encontrados, qtde_arquivos_pesquisados, qtde_documentos_nao_encontrados, qtde_documentos_encontrados

def criar_relatorio(documentos_encontrados, documentos_nao_encontrados, qtde_arquivos_pesquisados, nome_arquivo, diretorio_atual): #função responsável por criar relatório dos arquivos pesquisados
    workbook = openpyxl.Workbook() #declarando o workbook
    sheet = workbook.active #ativando o workbook do openpyxl

    #inicializando as colunas principais
    sheet['A1'] = 'Documentos em arquivos' 
    sheet['B1'] = 'Nomenclatura arquivo'
    sheet['C1'] = 'Não encontrado'

    #declarando as linhas dos documentos encontrados e não encontrados
    linha_docs_encontrados = 1
    linha_docs_nao_encontrados = 1

    #iniciando laço para percorrer os documentos encontrados
    for documento, arquivos in documentos_encontrados.items():
        linha_docs_encontrados += 1 #somando na variável para virar "A2" e sucessivamente.
        sheet.cell(row=linha_docs_encontrados, column=1, value=documento) #aplicando o documento encontrado na linha atual e coluna 1
        for i, arquivo in enumerate(arquivos, start=1):
            sheet.cell(row=linha_docs_encontrados, column=2, value=arquivo) #aplicando o arquivo encontrado na linha atual e coluna 1

    #laço que percorre os documentos não encontrados
    for documento in documentos_nao_encontrados:
        linha_docs_nao_encontrados += 1 #somando na variável para virar "C2" e sucessivamente.
        sheet.cell(row=linha_docs_nao_encontrados, column=3).value = documento #aplicando o documento não encontrado na linha atual e coluna 3

    #mudando a aplicação para o diretório escolhido pelo usuário
    os.chdir(diretorio_atual)
    nome_arquivo = incrementar_nome_arquivo(nome_arquivo) #chamando a função que incrementa na nomenclatura do arquivo
    
    caminho_relatorio = os.path.join(diretorio_atual, nome_arquivo) #juntando o diretório com o nome do arquivo para salvar o mesmo 

    #tentando salvar o relatório, se estiver aberto será retorno uma except
    try:
        workbook.save(caminho_relatorio)
        messagebox.showinfo("Relatório Gerado", f"Relatório gerado com sucesso: {nome_arquivo}")
    except PermissionError:
        messagebox.showerror("Erro", "Não foi possível gerar o relatório, pois está aberto por outro programa.")

def incrementar_nome_arquivo(nome_arquivo): #função responsável por iterar no salvamento de um novo arquivo para diferentes relatórios
    count = 1
    while os.path.exists(nome_arquivo): #enquanto a nomenclatura atual existir, será adicionado uma numeração única
        nome_arquivo = f"RelatorioProcessamento({count}).xlsx"
        count += 1
    return nome_arquivo

def btn_selecionar_diretorio(entry): #função responsável por dar a entrada de um diretório
    global diretorio_selecionado
    diretorio_selecionado = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, diretorio_selecionado)

def btn_pesquisar():
    inicio = time.time()  # Inicia a contagem do tempo no começo da função

    diretorio_atual = diretorio_selecionado  # Diretório selecionado pelo usuário
    if not diretorio_atual:
        messagebox.showerror("Erro", "Por favor, selecione um diretório.")  # Erro se não for selecionado um diretório
        return

    # Declarando as extensões que não vão ser validadas
    extensoes_descartadas = (".fpl", ".zip", ".ini", ".pdf", ".xlsx")

    # Se o botão de validar o fpl estiver ativo, será retirado ".fpl" das extensões descartadas
    if var_fpl.get() == 1:
        extensoes_descartadas = extensoes_descartadas[1:]

    # Declarando nomenclaturas que não devem ser percorridas
    validacao_nomenclatura = []

    # Capturando os documentos da função carregar documentos
    documentos = carregar_documentos(documentos_textbox)

    # Chamando a função pesquisar_documentos para fazer a pesquisa
    documentos_encontrados, documentos_nao_encontrados, qtde_arquivos_pesquisados, qtde_documentos_nao_encontrados, qtde_documentos_encontrados = pesquisar_documentos(
        diretorio_atual, extensoes_descartadas, validacao_nomenclatura, documentos)

    # Chamando a função para criar o relatório
    criar_relatorio(documentos_encontrados, documentos_nao_encontrados, qtde_arquivos_pesquisados, "RelatorioProcessamento.xlsx", diretorio_atual)

    # Inserindo os dados
    tempo_execucao = time.time() - inicio  # Calcula o tempo de execução
    minutos, segundos = divmod(tempo_execucao, 60)  # Divide o tempo em minutos e segundos
    tempo_formatado = f'{int(minutos):02}:{int(segundos):02}'  # Formata o tempo no formato MM:SS

    log_textbox.insert(tk.END, f'Quantidade de arquivos pesquisados: {qtde_arquivos_pesquisados}\n')
    log_textbox.insert(tk.END, f'Total de documentos não encontrados em arquivo: {qtde_documentos_nao_encontrados}\n')
    log_textbox.insert(tk.END, f'Total de documentos encontrados em arquivo: {qtde_documentos_encontrados}\n')
    log_textbox.insert(tk.END, f'Tempo de execução: {tempo_formatado}\n\n')

def limpar_log(): #função responsável pelo botão de limpar o log
    log_textbox.delete('1.0', tk.END)

def criar_dataframe(file_path): #função responsável pela criação do dataframe
    #criando dataframe
    df = pd.read_excel(file_path)
    
    #aplicando filtro para que todas nomenclaturas fiquem em mínusculo
    df['Nome Arquivo'] = df['Nome Arquivo'].str.lower()

    df['Processo'] = df['Processo'].str.lower()

    #aplicando um filtro para retirar o que vai ser desnecessário para a validação. Por exemplo: retirando tudo que tem ret*
    df = df.loc[~df['Processo'].str.contains(r"redecard_envioarscliente_rem.*|ret.*|reprog.*|enviopos.*", na=False)]

    df['Nome Arquivo'] = df['Nome Arquivo'].str.replace('xlsx', '')

    qtde_registros_concatenados = 0
    qtde_registros_originais = 0

    if df['Nome Arquivo'].str.contains(r"concaten.*").any(): #utilizando método any() para caso haja qualquer arquivo concaten, entre nessa validação.
        #aplicando um filtro para pegar tudo que tem concaten, porém retirando o que tem .fpl, para validar se tudo que foi concatenado bate com os originais
        df_registros_concatenados = df[df['Nome Arquivo'].str.contains(r"concaten.*") & ~df['Nome Arquivo'].str.contains(r'\.fpl')]
        
        qtde_registros_concatenados = df_registros_concatenados['Registros'].sum() #faz a soma somente dos registros concatenados
        
        print(f"Quantidade registros concatenados: {qtde_registros_concatenados}")

        #df_registros_originais = df[df['Nome Arquivo'].str.contains(r'import_padrao.*')]
        #aplicando um filtro para pegar somente os arquivos originais padrão API.
        df_registros_originais = df[df['Status'].str.contains('Concatenado')]

        qtde_registros_originais = df_registros_originais['Registros'].sum() #realizando a soma dos registros originais.
        
        print(f"Quantidade registros originais: {qtde_registros_originais}")

    #classificando os arquivos pela nomenclatura de A a Z
    df = df.sort_values(by='Nome Arquivo', ascending=True)


    #df['Registros JALL'] = df['Registros JALL'].apply(lambda x: int(x) if pd.notnull(x) else x)

    #verificando caso o primeiro registro termine com .fpl, só comece a validação após não houver mais o mesmo.
    while not df.empty and df.iloc[0]['Nome Arquivo'].endswith('.fpl'):
            df = df.iloc[1:] #pula a linha atual

    print(df)

    return df, qtde_registros_concatenados, qtde_registros_originais

def validar_quantidades(dataframe): #função responsável por validar as quantidades
    divergenciasPlanilha = []
    arquivosValidados = []

    #iniciando laço de repetição para percorrer todo o dataframe do processamento
    for i in range(0, len(dataframe)):
        original = dataframe.iloc[i] #declarando variável do arquivo original
        hora = dataframe.iloc[i] #declarando variável para pegar a data e hora do processamento
        if i < len(dataframe) - 1: #validação para acessar os arquivos de validação corretamente e não exceder o len da planilha
            validacao = dataframe.iloc[i + 1]
        
        #validando se a nomenclatura do arquivo original NÃO termina com fpl e se o arquivo de validação (subsequente) termina com .fpl
        if not original['Nome Arquivo'].endswith('.fpl') and validacao['Nome Arquivo'].endswith('.fpl'):
            if original['Registros'] != validacao['Registros']: #verificando se há divergência de valor
                divergenciasPlanilha.append({ #inserindo os valores dentro da lista declarada para montagem da planilha de divergência
                    'Nome Original': original['Nome Arquivo'],
                    'Quantidade Original': original['Registros'],
                    'Nome Validacao': validacao['Nome Arquivo'],
                    'Quantidade Validacao': validacao['Registros'],
                    'Hora Processamento': hora['Data']
                })

            else:
                 arquivosValidados.append({ #Caso não haja divergência, armazenando os valores validados da outra lista declarada
                      'NomeArquivo': original['Nome Arquivo'],
                      'Quantidade': original['Registros']
                 })
                 arquivosValidados.append({
                      'NomeArquivo': validacao['Nome Arquivo'],
                      'Quantidade': validacao['Registros']
                 })
        #Validação para qualquer arquivo que esteja no status entregar ou erro, ser adicionado a planilha de divergência para análise
        if original['Status'] == "Entregar" or original['Status'] == "Erro":
             divergenciasPlanilha.append({
                    'Nome Original': original['Nome Arquivo'],
                    'Quantidade Original': original['Registros'],
                    'Nome Validacao': original['Nome Arquivo'],
                    'Quantidade Validacao': original['Registros'],
                    'Hora Processamento': hora['Data'],
                    'Status': original['Status']
                })

    #VERIFICAR     
    divergenciasPlanilhaConvertida = "\n\n".join(
    [f"{item.get('Nome Original', 'N/A')} / {item.get('Quantidade Original', 'N/A')}\n"
     f"{item.get('Nome Validacao', 'Não disponível')} / {item.get('Quantidade Validacao', 'N/A')}\n"
     f"Horário: {item.get('Hora Processamento', 'N/A')}".strip() for item in divergenciasPlanilha]
)

    return pd.DataFrame(divergenciasPlanilha), pd.DataFrame(arquivosValidados), divergenciasPlanilhaConvertida

def concatenar_arquivos(caminhoArquivos, caminhoRelatorio): #função responsável por contenar arquivos dentro do diretório
    listarArquivos = os.listdir(caminhoArquivos) #listando os arquivos no diretório
    
    #listando os arquivos que começam com "Arquivos Processados"
    listarCaminhoEArquivos = [caminhoArquivos + '\\' + arquivo for arquivo in listarArquivos if arquivo.startswith('Arquivos Processados')]

    dfConcatenado = pd.DataFrame()

    for arquivo in listarCaminhoEArquivos:
        dados = pd.read_excel(arquivo)
        dfConcatenado = dfConcatenado._append(dados)

    dfConcatenado.to_excel(caminhoRelatorio)

    return dfConcatenado

def btn_validar():
    diretorio_atual = diretorio_selecionado
    if not diretorio_atual:
        messagebox.showerror("Erro", "Por favor, selecione um diretório.")
        return
    
    caminhoRelatorio = f"{diretorio_atual}\RelatorioConcatenado.xlsx"

    listaArquivosDiretorio = os.listdir(diretorio_atual)

    listaArquivosProcessados = [arquivo for arquivo in listaArquivosDiretorio if arquivo.startswith("Arquivos Processados")]

    nomenclaturaArquivoProcessado = listaArquivosProcessados[0]

    planilhaOrdenada = f'{diretorio_atual}/RelatorioOrdenado(VERIFICAR).xlsx'

    if len(listaArquivosProcessados) > 1:
        dfConcatenado = concatenar_arquivos(diretorio_atual, caminhoRelatorio)

        dfConcatenado, qtde_registros_concatenados, qtde_registros_Originais = criar_dataframe(f"{diretorio_atual}/RelatorioConcatenado.xlsx")

        dfConcatenado.to_excel(planilhaOrdenada, index=False)

        divergenciasPlanilha, arquivosValidados, divergenciasPlanilhaConvertida = validar_quantidades(dfConcatenado)    
        
        if qtde_registros_concatenados and qtde_registros_Originais != 0 and divergenciasPlanilha.empty:
            if qtde_registros_concatenados == qtde_registros_Originais:
                log_textbox2.insert(tk.END,"Foi identificado fluxo de arquivos originais x contenados.\n"
                                            f"Qtde Arquivos originais: {qtde_registros_Originais}\n"
                                            f"Quantidade arquivos concatenados: {qtde_registros_concatenados}\n\n"
                                            "Não foi encontrado divergências dentre os arquivos validados. Confira o relatorio gerado 'RelatorioOrdenado(VERIFICAR).xlsx' para validação."
                                    )
            else:
                log_textbox2.insert(tk.END,"Foi identificado fluxo de arquivos originais x contenados.\n"
                                        f"Qtde Arquivos originais: {qtde_registros_Originais}\n"
                                        f"Quantidade arquivos concatenados: {qtde_registros_concatenados}\n\n"
                                        "Consulte as planilhas para analisar a diferença dentre original e concatenados"
                                    )
        elif qtde_registros_concatenados and qtde_registros_Originais != 0 and not divergenciasPlanilha.empty:
            log_textbox2.insert(tk.END,"Foi identificado fluxo de arquivos originais x contenados.\n"
                                        f"Qtde Arquivos originais: {qtde_registros_Originais}\n"
                                        f"Quantidade arquivos concatenados: {qtde_registros_concatenados}\n \n"
                                        "Foram encontradas as seguintes divergências entre arquivos abaixo:\n\n"
                                        f"{divergenciasPlanilhaConvertida}\n\n"
                                        "Consulte o relatorio de divergência para mais informações."
                                )  
        elif not divergenciasPlanilha.empty:
            log_textbox2.insert(tk.END,f'Foram encontradas as seguintes divergências entre arquivos abaixo:\n \n{divergenciasPlanilhaConvertida}\n \nConsulte o relatorio de divergencia para mais informações.')

        else:
            log_textbox2.insert(tk.END, f'Não foi encontrado divergências dentre os arquivos validados.\n' 
                                        "Confira o relatorio gerado 'RelatorioOrdenado(VERIFICAR).xlsx' para validação."
                                )

    else:
        df, qtde_registros_concatenados, qtde_registros_Originais = criar_dataframe(f"{diretorio_atual}/{nomenclaturaArquivoProcessado}")
        
        df.to_excel(planilhaOrdenada, index=False)
        
        divergenciasPlanilha, arquivosValidados, divergenciasPlanilhaConvertida = validar_quantidades(df)

        if qtde_registros_concatenados and qtde_registros_Originais != 0 and divergenciasPlanilha.empty:
            if qtde_registros_concatenados == qtde_registros_Originais:
                log_textbox2.insert(tk.END,"Foi identificado fluxo de arquivos originais x contenados.\n"
                                            f"Qtde Arquivos originais: {qtde_registros_Originais}\n"
                                            f"Quantidade arquivos concatenados: {qtde_registros_concatenados}\n\n"
                                            "Não foi encontrado divergências dentre os arquivos validados. Confira o relatorio gerada 'RelatorioOrdenado(VERIFICAR).xlsx' para validação."
                                    )
            else:
                log_textbox2.insert(tk.END,"Foi identificado fluxo de arquivos originais x contenados.\n"
                                        f"Qtde Arquivos originais: {qtde_registros_Originais}\n"
                                        f"Quantidade arquivos concatenados: {qtde_registros_concatenados}\n\n"
                                        "Consulte as planilhas para analisar a diferença dentre original e concatenados"
                                    )
        elif qtde_registros_concatenados and qtde_registros_Originais != 0 and not divergenciasPlanilha.empty:
            log_textbox2.insert(tk.END,"Foi identificado fluxo de arquivos originais x contenados.\n"
                                        f"Qtde Arquivos originais: {qtde_registros_Originais}\n"
                                        f"Quantidade arquivos concatenados: {qtde_registros_concatenados}\n \n"
                                        "Foram encontradas as seguintes divergências entre arquivos abaixo:\n\n"
                                        f"{divergenciasPlanilhaConvertida}\n\n"
                                        "Consulte o relatorio de divergência para mais informações."
                                ) 
        elif not divergenciasPlanilha.empty:
            log_textbox2.insert(tk.END,f'Foram encontradas as seguintes divergências entre arquivos abaixo:\n \n{divergenciasPlanilhaConvertida}\n \nConsulte a planilha RelatorioDivergencias.xlsx para mais informações.')

        else:
            log_textbox2.insert(tk.END, f'Não foi encontrado nenhuma divergência dentre os arquivos validados. Consulte o relatório RelatorioOrdenado(VERIFICAR).xlsx para verificação.\n')


    planilhaDivergencias = f'{diretorio_atual}/RelatorioDivergencias.xlsx'
    #planilhaPosValidacao = r'c:\Users\gabriel.ferreira\Downloads\ArquivosValidados.xlsx'
    divergenciasPlanilha.to_excel(planilhaDivergencias, index=False)
    #arquivosValidados.to_excel(planilhaPosValidacao, index=False)
    
def busca_reversa(diretorio_atual, posicao, comprimento, nomenclaturaDescartada):
    documentosExtraidos = []
    try:
        posicao = int(posicao)
        comprimento = int(comprimento)

        for root, _, files in os.walk(diretorio_atual):
            for file in files:
                if not file.endswith(nomenclaturaDescartada):
                    caminho_arquivo = os.path.join(root, file)

                    with open(caminho_arquivo, 'r', encoding='utf-8', errors="ignore") as f:
                        for linha in f:
                            novaPosicao = posicao - 1 if posicao > 1 else 0
                            documento = linha[novaPosicao:novaPosicao + comprimento]
                            documentosExtraidos.append(documento)
        
        return documentosExtraidos
    except ValueError:
        return ["Erro: Posição e comprimento precisam ser números inteiros."]

def btn_pesquisa_reversa():
    diretorio_atual = diretorio_selecionado
    if not diretorio_atual:
        messagebox.showerror("Erro", "Por favor, selecione um diretório.")
        return
    
    extensoes_descartadas = (".fpl", ".zip", ".ini", ".pdf", ".xlsx")
    
    posicao = entry_posicao.get()
    comprimento = entry_comprimento.get()
    log_text.delete(1.0, tk.END)  # Limpa o log
    documentosExtraidos = busca_reversa(diretorio_atual, posicao, comprimento, extensoes_descartadas)
    
    for doc in documentosExtraidos:
        log_text.insert(tk.END, doc + "\n") 

# Criar a janela
janela = tk.Tk()
janela.title("Pesquisa de Documentos")
janela.geometry("600x600")  # Aumentar o tamanho da janela para acomodar o Notebook
janela.resizable(False,False)

# Aplicar estilo do ttkbootstrap
style = Style(theme='flatly')

# Criar o Notebook
notebook = ttk.Notebook(janela)
notebook.pack(pady=10, expand=True, fill='both')

# Criar a primeira aba
aba1 = ttk.Frame(notebook)
notebook.add(aba1, text='Pesquisar Documentos')

# Adicionar widgets à primeira aba
diretorio_frame = tk.Frame(aba1)
diretorio_frame.pack(pady=20)

diretorio_label = tk.Label(diretorio_frame, text="Diretório:", font=("Montserrat", 8, 'bold'))
diretorio_label.grid(row=0, column=0, padx=5)

entry_diretorio_aba1 = tk.Entry(diretorio_frame, width=50)
entry_diretorio_aba1.grid(row=0, column=1, padx=5)

selecionar_diretorio_button = tk.Button(diretorio_frame, text="Selecionar", width=10, command=lambda: btn_selecionar_diretorio(entry_diretorio_aba1))
selecionar_diretorio_button.grid(row=0, column=2, padx=5)

documentos_label = tk.Label(aba1, text="Insira os documentos (um embaixo do outro)", font=("Montserrat", 11, 'bold'))
documentos_label.pack(pady=5)

documentos_textbox = scrolledtext.ScrolledText(aba1, width=50, height=10)
documentos_textbox.pack()

pesquisar_button = tk.Button(
    aba1, 
    text="Pesquisar", 
    width=11, 
    command=btn_pesquisar
)
pesquisar_button.pack(pady=5, padx=5)

var_fpl = tk.IntVar()
checkBoxFpl = tk.Checkbutton(aba1, text='Pesquisar em arquivos .FPL', font=("Montserrat", 8, 'bold'), variable=var_fpl, onvalue=1, offvalue=0)
checkBoxFpl.pack(pady=5)

log_frame = tk.Frame(aba1)
log_frame.pack(padx=5, pady=5)

log_textbox = scrolledtext.ScrolledText(log_frame, width=60, height=10)
log_textbox.pack()

limpar_log_button = tk.Button(log_frame, text="Limpar Log", width=10, command=limpar_log)
limpar_log_button.pack(pady=5)

#############################################################################################################################################
##############################################################################################################################################

aba2 = ttk.Frame(notebook)
notebook.add(aba2, text='Validação Schedulle')

# Configurar colunas
aba2.grid_columnconfigure(0, weight=1)
aba2.grid_columnconfigure(1, weight=3)  # Aumentar a largura da coluna do Entry
aba2.grid_columnconfigure(2, weight=1)

# Título centralizado
nova_funcionalidade_label = tk.Label(aba2, text="Validação Schedulle", font=("Montserrat", 12, 'bold'))
nova_funcionalidade_label.grid(row=0, column=0, columnspan=3, pady=10, padx=3, sticky="nsew")

label_descricao_aba2 = tk.Label(
    aba2, 
    text="Selecione o diretório com os relatórios extraídos do ga, para que seja extraído os dados, validados e gerados planilhas ordenadas para análise e validação.",
    wraplength=575,  # Controla a quebra de linha
    justify="left",  # Alinhamento do texto
    font=("Roboto", 10),  # Estilo da fonte
    bg="#f0f0f0",  # Cor de fundo suave
    fg="#333333",  # Cor do texto mais escura
)
label_descricao_aba2.grid(row=1, column=0, columnspan=3, pady=3, padx=2, sticky="n")

# Label e Entry do Diretório
label_diretorio_aba2 = tk.Label(aba2, text="Diretório:", font=("Montserrat", 8, 'bold'))
label_diretorio_aba2.grid(row=2, column=0, padx=2, pady=10, sticky="e")

entry_diretorio_aba2 = tk.Entry(aba2, width=50)
entry_diretorio_aba2.grid(row=2, column=1, padx=5, pady=10, sticky="we")  # Expandir o Entry horizontalmente

button_selecionar_diretorio_aba2 = tk.Button(aba2, text="Selecionar", width=10, command=lambda: btn_selecionar_diretorio(entry_diretorio_aba2))
button_selecionar_diretorio_aba2.grid(row=2, column=2, padx=5, pady=10, sticky="w")  # Alinhar o botão à esquerda

log_textbox2 = scrolledtext.ScrolledText(aba2, width=90, height=20)
log_textbox2.grid(row=3, column=0, columnspan=3, padx=10, pady=10,sticky="n")

# Botão de Validar centralizado
validar_button2 = tk.Button(aba2, text="Validar", width=11, command=btn_validar)
validar_button2.grid(row=4, column=0, columnspan=3, pady=10)

#############################################################################################################################################
#############################################################################################################################################

aba3 = ttk.Frame(notebook)
notebook.add(aba3, text='Pesquisa Reversa')

# Configurando as colunas para expansão
#aba3.grid_columnconfigure(0, weight=1)
#aba3.grid_columnconfigure(1, weight=1)

# Título principal
label_aba3 = tk.Label(aba3, text="Pesquisa Reversa", font=("Montserrat", 12, 'bold'))
label_aba3.grid(row=0, column=0, columnspan=3, pady=10, padx=2, sticky="n")  # Centralizado no topo

# Descrição da funcionalidade
label_descricao_aba3 = tk.Label(
    aba3, 
    text="Declare a posição dentro dos arquivos que deseja pesquisar e o comprimento da sequência de caracteres, para que seja pesquisado e extraído de arquivo a arquivo no seu diretório selecionado.",
    wraplength=580,  # Ajustar o comprimento para manter o texto dentro da área visível
    justify="left",  # Centraliza o texto horizontalmente
    font=("Roboto", 10),  # Estilo da fonte
    bg="#f0f0f0",  # Cor de fundo suave
    fg="#333333",  # Cor do texto mais escura
)
label_descricao_aba3.grid(row=1, column=0, columnspan=3, pady=3, padx=15, sticky="n")  # Alinhado e centralizado

# Diretório
label_diretorio_aba3 = tk.Label(aba3, text="Diretório:", font=("Montserrat", 8, 'bold'))
label_diretorio_aba3.grid(row=2, column=0, padx=2, pady=10, sticky="e")  # Alinhado à direita

entry_diretorio_aba3 = tk.Entry(aba3, width=50)
entry_diretorio_aba3.grid(row=2, column=1, padx=2, pady=10, sticky="we")  # Expandir o Entry horizontalmente

button_selecionar_diretorio_aba3 = tk.Button(aba3, text="Selecionar", width=10, command=lambda: btn_selecionar_diretorio(entry_diretorio_aba3))
button_selecionar_diretorio_aba3.grid(row=2, column=2, padx=5, pady=10, sticky="w")  # Alinhado à esquerda

# Posição
tk.Label(aba3, text="Posição:", font=("Montserrat", 8, 'bold')).grid(row=3, column=0, padx=5, pady=2, sticky="e")  # Alinhado à direita
entry_posicao = tk.Entry(aba3, width=8)
entry_posicao.grid(row=3, column=1, padx=2, pady=5, sticky="w")  # Alinhado à esquerda

# Comprimento
tk.Label(aba3, text="Comprimento:", font=("Montserrat", 8, 'bold')).grid(row=4, column=0, padx=5, pady=5, sticky="e")  # Alinhado à direita
entry_comprimento = tk.Entry(aba3, width=8)
entry_comprimento.grid(row=4, column=1, padx=2, pady=5, sticky="w")  # Alinhado à esquerda

# Botão de pesquisa centralizado
pesquisa_reversa_button = tk.Button(aba3, text="Pesquisar", width=12, command=btn_pesquisa_reversa)
pesquisa_reversa_button.grid(row=5, column=0, columnspan=3, pady=10, sticky="n")  # Centralizado

# Label de informações
pesquisa_reversa_label = tk.Label(aba3, text="Informações extraídas dos arquivos pesquisados", font=("Montserrat", 11, 'bold'))
pesquisa_reversa_label.grid(row=6, column=0, columnspan=3, pady=2, sticky="n")  # Centralizado

# Log de texto para informações de extração
log_text = scrolledtext.ScrolledText(aba3, width=60, height=15)
log_text.grid(row=7, column=0, columnspan=3, padx=10, pady=10, sticky="n")  # Centralizado com espaço nas laterais

# Iniciar a janela
inicio = time.time()
janela.mainloop()